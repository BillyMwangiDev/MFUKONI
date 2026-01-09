"""
Views for Mfukoni tracker application.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.urls import reverse
from .db_manager import get_db
from .forms import TransactionForm, FilterForm, CategoryForm, BudgetForm, DateRangeForm
from my_rdbms.exceptions import DatabaseError, ConstraintError


def dashboard(request):
    """Display financial dashboard."""
    try:
        db = get_db()
        summary = db.get_summary()
        recent = db.get_all_transactions(limit=10)
        spending_by_category = db.get_spending_by_category()
        
        # Get category names for recent transactions
        categories = {c['id']: c['name'] for c in db.get_all_categories()}
        for trans in recent:
            trans['category_name'] = categories.get(trans.get('category_id'), 'Unknown')
        
        context = {
            'summary': summary,
            'recent_transactions': recent,
            'spending_by_category': spending_by_category,
        }
        return render(request, 'tracker/dashboard.html', context)
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        # Return empty context on error
        context = {
            'summary': {'total_income': 0.0, 'total_expenses': 0.0, 'balance': 0.0, 'transaction_count': 0},
            'recent_transactions': [],
            'spending_by_category': [],
        }
        return render(request, 'tracker/dashboard.html', context)


def transaction_list(request):
    """Display list of all transactions with filtering."""
    try:
        db = get_db()
        form = FilterForm(request.GET)
        date_form = DateRangeForm(request.GET)
        
        category_id = None
        trans_type = None
        search_query = None
        start_date = None
        end_date = None
        
        if form.is_valid():
            # category_id comes as string from ChoiceField, convert to int or None
            cat_id_str = form.cleaned_data.get('category_id')
            category_id = int(cat_id_str) if cat_id_str and cat_id_str != '' else None
            trans_type = form.cleaned_data.get('trans_type') or None
            search_query = form.cleaned_data.get('search') or None
        
        if date_form.is_valid():
            if date_form.cleaned_data.get('start_date'):
                start_date = str(date_form.cleaned_data['start_date'])
            if date_form.cleaned_data.get('end_date'):
                end_date = str(date_form.cleaned_data['end_date'])
        
        # Get transactions
        try:
            if start_date and end_date:
                transactions = db.get_transactions_by_date_range(start_date, end_date)
                # Apply additional filters
                if category_id:
                    transactions = [t for t in transactions if t.get('category_id') == category_id]
                if trans_type:
                    transactions = [t for t in transactions if t.get('type') == trans_type]
            elif search_query:
                transactions = db.search_transactions(search_query)
            else:
                transactions = db.get_all_transactions(category_id=category_id, trans_type=trans_type)
        except Exception as e:
            messages.warning(request, f'Error loading transactions: {str(e)}')
            transactions = []
        
        # Get category names
        try:
            categories = {c['id']: c['name'] for c in db.get_all_categories()}
            for trans in transactions:
                trans['category_name'] = categories.get(trans.get('category_id'), 'Unknown')
        except Exception:
            # If categories fail to load, just mark as Unknown
            for trans in transactions:
                trans['category_name'] = 'Unknown'
        
        context = {
            'transactions': transactions,
            'form': form,
            'date_form': date_form,
        }
        return render(request, 'tracker/transaction_list.html', context)
    except Exception as e:
        messages.error(request, f'Error loading transaction list: {str(e)}')
        context = {
            'transactions': [],
            'form': FilterForm(),
            'date_form': DateRangeForm(),
        }
        return render(request, 'tracker/transaction_list.html', context)


def add_transaction(request):
    """Add a new transaction."""
    db = get_db()
    categories = db.get_all_categories()
    
    # Check if user has categories
    if not categories:
        messages.warning(
            request, 
            'Please create at least one category before adding transactions. '
            '<a href="{}" class="alert-link">Create Category</a>'.format(
                reverse('tracker:category_list')
            )
        )
        return redirect('tracker:category_list')
    
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            try:
                db = get_db()
                db.add_transaction(
                    category_id=int(form.cleaned_data['category_id']),
                    amount=float(form.cleaned_data['amount']),
                    description=form.cleaned_data['description'],
                    date=str(form.cleaned_data['date']),
                    trans_type=form.cleaned_data['trans_type']
                )
                # Get updated balance
                summary = db.get_summary()
                messages.success(
                    request, 
                    f'Transaction added successfully! Current balance: KES {summary["balance"]:,.2f}'
                )
                return redirect('tracker:dashboard')
            except (DatabaseError, ConstraintError) as e:
                messages.error(request, f'Error saving transaction: {str(e)}')
        else:
            # Form validation failed - display specific errors
            # Also log for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Form errors: {form.errors}")
            logger.error(f"POST data: {request.POST}")
            
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        form = TransactionForm()
        # Pre-fill from query parameters (for quick add)
        trans_type = request.GET.get('type')
        category_name = request.GET.get('category')
        if trans_type:
            form.fields['trans_type'].initial = trans_type
        if category_name:
            db = get_db()
            categories = db.get_all_categories()
            for cat in categories:
                if cat['name'] == category_name:
                    form.fields['category_id'].initial = str(cat['id'])  # Convert to string for ChoiceField
                    break
    
    context = {'form': form}
    return render(request, 'tracker/add_transaction.html', context)


def edit_transaction(request, transaction_id):
    """Edit an existing transaction."""
    db = get_db()
    transaction = db.get_transaction(transaction_id)
    
    if not transaction:
        messages.error(request, 'Transaction not found.')
        return redirect('tracker:transaction_list')
    
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            try:
                db.update_transaction(
                    transaction_id=transaction_id,
                    category_id=int(form.cleaned_data['category_id']),
                    amount=float(form.cleaned_data['amount']),
                    description=form.cleaned_data['description'],
                    date=str(form.cleaned_data['date']),
                    trans_type=form.cleaned_data['trans_type']
                )
                messages.success(request, 'Transaction updated successfully!')
                return redirect('tracker:transaction_list')
            except (DatabaseError, ConstraintError) as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        # Pre-populate form with existing data
        form = TransactionForm(initial={
            'trans_type': transaction.get('type'),
            'category_id': str(transaction.get('category_id')),  # Convert to string for ChoiceField
            'amount': transaction.get('amount'),
            'description': transaction.get('description'),
            'date': transaction.get('date'),
        })
    
    context = {
        'form': form,
        'transaction': transaction,
    }
    return render(request, 'tracker/edit_transaction.html', context)


def delete_transaction(request, transaction_id):
    """Delete a transaction."""
    if request.method == 'POST':
        try:
            db = get_db()
            db.delete_transaction(transaction_id)
            messages.success(request, 'Transaction deleted successfully!')
        except DatabaseError as e:
            messages.error(request, f'Error: {str(e)}')
    
    return redirect('tracker:transaction_list')


def category_list(request):
    """Display and manage categories."""
    db = get_db()
    categories = db.get_all_categories()
    
    # Separate by type
    income_categories = [c for c in categories if c.get('type') == 'income']
    expense_categories = [c for c in categories if c.get('type') == 'expense']
    
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            try:
                db.add_category(
                    name=form.cleaned_data['name'],
                    cat_type=form.cleaned_data['cat_type']
                )
                messages.success(request, 'Category added successfully!')
                return redirect('tracker:category_list')
            except (DatabaseError, ConstraintError) as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = CategoryForm()
    
    context = {
        'income_categories': income_categories,
        'expense_categories': expense_categories,
        'form': form,
        'has_categories': len(categories) > 0,
    }
    return render(request, 'tracker/category_list.html', context)


def edit_category(request, category_id):
    """Edit an existing category."""
    db = get_db()
    category = db.get_category(category_id)
    
    if not category:
        messages.error(request, 'Category not found.')
        return redirect('tracker:category_list')
    
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            try:
                db.update_category(
                    category_id=category_id,
                    name=form.cleaned_data['name'],
                    cat_type=form.cleaned_data['cat_type']
                )
                messages.success(request, 'Category updated successfully!')
                return redirect('tracker:category_list')
            except (DatabaseError, ConstraintError) as e:
                messages.error(request, f'Error: {str(e)}')
    else:
        form = CategoryForm(initial={
            'name': category.get('name'),
            'cat_type': category.get('type'),
        })
    
    context = {
        'form': form,
        'category': category,
    }
    return render(request, 'tracker/edit_category.html', context)


def delete_category(request, category_id):
    """Delete a category."""
    if request.method == 'POST':
        try:
            db = get_db()
            db.delete_category(category_id)
            messages.success(request, 'Category deleted successfully!')
        except DatabaseError as e:
            messages.error(request, f'Error: {str(e)}')
    
    return redirect('tracker:category_list')


# SQL console removed - CRUD operations automatically execute SQL queries


def get_categories_ajax(request):
    """AJAX endpoint to get categories by type."""
    try:
        trans_type = request.GET.get('type', 'income')
        db = get_db()
        categories = db.get_all_categories(cat_type=trans_type)
        # Ensure categories are in the correct format for JSON
        formatted_categories = [
            {'id': cat.get('id'), 'name': cat.get('name')}
            for cat in categories
        ]
        return JsonResponse({'categories': formatted_categories})
    except Exception as e:
        return JsonResponse({'error': str(e), 'categories': []}, status=500)


def budget_list(request):
    """Display and manage budgets."""
    db = get_db()
    
    # Get current month
    from datetime import datetime
    current_month = datetime.now().strftime("%Y-%m")
    
    # Get budget status for current month
    budget_status = db.get_budget_status(current_month)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            try:
                # Check if budget already exists for this category and month
                # category_id is a string from ChoiceField, convert to int
                existing = db.get_budget_for_category(
                    int(form.cleaned_data['category_id']),
                    form.cleaned_data['month']
                )
                if existing:
                    # Update existing budget
                    db.update_budget(existing['id'], float(form.cleaned_data['monthly_limit']))
                    messages.success(request, 'Budget updated successfully!')
                else:
                    # Create new budget
                    db.set_budget(
                        category_id=int(form.cleaned_data['category_id']),
                        monthly_limit=float(form.cleaned_data['monthly_limit']),
                        month=form.cleaned_data['month']
                    )
                    messages.success(request, 'Budget set successfully!')
                return redirect('tracker:budget_list')
            except (DatabaseError, ConstraintError) as e:
                messages.error(request, f'Error: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors in the form below.')
    else:
        form = BudgetForm()
    
    context = {
        'budget_status': budget_status,
        'form': form,
        'current_month': current_month,
    }
    return render(request, 'tracker/budget_list.html', context)


def delete_budget(request, budget_id):
    """Delete a budget."""
    if request.method == 'POST':
        try:
            db = get_db()
            db.delete_budget(budget_id)
            messages.success(request, 'Budget deleted successfully!')
        except DatabaseError as e:
            messages.error(request, f'Error: {str(e)}')
    
    return redirect('tracker:budget_list')


def reports(request):
    """Financial reports and analytics."""
    from datetime import datetime
    import calendar
    
    try:
        db = get_db()
        now = datetime.now()
        
        try:
            current_summary = db.get_monthly_summary(now.year, now.month)
            # Add month name
            current_summary['month_name'] = calendar.month_abbr[now.month]
        except (DatabaseError, Exception) as e:
            messages.error(request, f'Error loading current month summary: {str(e)}')
            current_summary = {
                'year': now.year,
                'month': now.month,
                'month_name': calendar.month_abbr[now.month],
                'income': 0.0,
                'expenses': 0.0,
                'balance': 0.0,
                'transaction_count': 0
            }
        
        # Get last 6 months summaries
        monthly_summaries = []
        for i in range(6):
            month = now.month - i
            year = now.year
            if month <= 0:
                month += 12
                year -= 1
            try:
                summary = db.get_monthly_summary(year, month)
                # Add month name
                summary['month_name'] = calendar.month_abbr[month]
                monthly_summaries.append(summary)
            except (DatabaseError, Exception) as e:
                # If one month fails, add empty summary
                monthly_summaries.append({
                    'year': year,
                    'month': month,
                    'month_name': calendar.month_abbr[month],
                    'income': 0.0,
                    'expenses': 0.0,
                    'balance': 0.0,
                    'transaction_count': 0
                })
        
        # Get top spending categories (all time)
        try:
            spending_by_category = db.get_spending_by_category()
            top_categories = spending_by_category[:5] if len(spending_by_category) > 5 else spending_by_category
        except (DatabaseError, Exception) as e:
            messages.warning(request, f'Error loading spending categories: {str(e)}')
            top_categories = []
        
        # Get budget status for current month to show in reports
        try:
            current_month_str = f"{now.year}-{now.month:02d}"
            budget_status = db.get_budget_status(current_month_str)
        except (DatabaseError, Exception) as e:
            messages.warning(request, f'Error loading budget status: {str(e)}')
            budget_status = []
        
        context = {
            'current_summary': current_summary,
            'monthly_summaries': monthly_summaries,
            'top_categories': top_categories,
            'budget_status': budget_status,
            'current_month': current_month_str,
        }
        return render(request, 'tracker/reports.html', context)
    except Exception as e:
        messages.error(request, f'An unexpected error occurred: {str(e)}')
        # Return minimal context to prevent template errors
        import calendar
        now = datetime.now()
        current_month_str = f"{now.year}-{now.month:02d}"
        context = {
            'current_summary': {
                'year': now.year,
                'month': now.month,
                'month_name': calendar.month_abbr[now.month],
                'income': 0.0,
                'expenses': 0.0,
                'balance': 0.0,
                'transaction_count': 0
            },
            'monthly_summaries': [],
            'top_categories': [],
            'budget_status': [],
            'current_month': current_month_str,
        }
        return render(request, 'tracker/reports.html', context)


def export_transactions(request):
    """Export transactions in selected format (CSV, PDF, or Excel)."""
    format_type = request.GET.get('format', 'csv').lower()
    
    try:
        db = get_db()
        transactions = db.get_all_transactions()
        
        # Get category names
        categories = {c['id']: c['name'] for c in db.get_all_categories()}
        
        # Prepare data
        data = []
        for trans in transactions:
            data.append({
                'date': trans.get('date', ''),
                'type': trans.get('type', '').title(),
                'category': categories.get(trans.get('category_id'), 'Unknown'),
                'description': trans.get('description', ''),
                'amount': float(trans.get('amount', 0))
            })
        
        if format_type == 'pdf':
            return _export_transactions_pdf(data)
        elif format_type == 'excel':
            return _export_transactions_excel(data)
        else:  # CSV (default)
            return _export_transactions_csv(data)
    except Exception as e:
        messages.error(request, f'Error exporting transactions: {str(e)}')
        return redirect('tracker:transaction_list')


def _export_transactions_csv(data):
    """Export transactions to CSV."""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="mfukoni_transactions.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Type', 'Category', 'Description', 'Amount (KES)'])
    
    for row in data:
        writer.writerow([
            row['date'],
            row['type'],
            row['category'],
            row['description'],
            row['amount']
        ])
    
    return response


def _export_transactions_pdf(data):
    """Export transactions to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2497F9'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph("<b>Mfukoni Finance Tracker - Transactions</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Prepare table data
    table_data = [['Date', 'Type', 'Category', 'Description', 'Amount (KES)']]
    
    total_amount = 0
    for row in data:
        table_data.append([
            row['date'],
            row['type'],
            row['category'],
            row['description'] or '-',
            f"{row['amount']:,.2f}"
        ])
        if row['type'].lower() == 'income':
            total_amount += row['amount']
        else:
            total_amount -= row['amount']
    
    # Add summary row
    table_data.append(['', '', '', '<b>Total Balance</b>', f"<b>{total_amount:,.2f}</b>"])
    
    # Create table
    table = Table(table_data, colWidths=[1*inch, 0.8*inch, 1*inch, 2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2497F9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F4FD')),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mfukoni_transactions.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _export_transactions_excel(data):
    """Export transactions to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Header row
    headers = ['Date', 'Type', 'Category', 'Description', 'Amount (KES)']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="2497F9", end_color="2497F9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Header row with smooth outer border
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        left_style = 'medium' if col_idx == 1 else 'thin'
        right_style = 'medium' if col_idx == 5 else 'thin'
        cell.border = Border(
            left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
            right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
            top=Side(style='medium', color='CCCCCC'),
            bottom=Side(style='thin', color='E0E0E0')
        )
    
    # Data rows
    total_amount = 0
    for row in data:
        ws.append([
            row['date'],
            row['type'],
            row['category'],
            row['description'] or '',
            row['amount']
        ])
        if row['type'].lower() == 'income':
            total_amount += row['amount']
        else:
            total_amount -= row['amount']
    
    # Add summary row
    ws.append(['', '', '', 'Total Balance', total_amount])
    summary_row = ws[ws.max_row]
    summary_fill = PatternFill(fill_type=None)  # No fill - transparent background
    for col_idx, cell in enumerate(summary_row, start=1):
        cell.fill = summary_fill
        cell.font = Font(bold=True, color="000000", size=11)  # Black, bold, visible font
        left_style = 'medium' if col_idx == 1 else 'thin'
        right_style = 'medium' if col_idx == 5 else 'thin'
        cell.border = Border(
            left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
            right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='medium', color='CCCCCC')
        )
    
    # Apply smooth borders and visible fonts to all data cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White background
    date_font = Font(bold=True, color="000000", size=11)  # Black, bold, visible font for dates
    data_font = Font(color="000000", size=11)  # Black font for all other data cells
    max_row = ws.max_row - 1  # Exclude summary row (already styled)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            # Explicitly set white fill to ensure visibility
            cell.fill = white_fill
            
            # Apply visible black font to all cells
            if col_idx == 1:  # Date column
                cell.font = date_font
                cell.number_format = '@'  # Format as text to prevent Excel date formatting
            else:
                cell.font = data_font
            
            left_style = 'medium' if col_idx == 1 else 'thin'
            right_style = 'medium' if col_idx == 5 else 'thin'
            top_style = 'thin'
            bottom_style = 'thin'
            
            cell.border = Border(
                left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                top=Side(style=top_style, color='E0E0E0'),
                bottom=Side(style=bottom_style, color='E0E0E0')
            )
            
            if col_idx == 5:  # Amount column
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column widths
    column_widths = {'A': 12, 'B': 10, 'C': 15, 'D': 30, 'E': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mfukoni_transactions.xlsx"'
    return response


def export_reports(request):
    """Export reports in selected format (PDF or Excel)."""
    format_type = request.GET.get('format', 'pdf').lower()
    
    try:
        db = get_db()
        from datetime import datetime
        
        now = datetime.now()
        
        # Get current month summary
        try:
            current_summary = db.get_monthly_summary(now.year, now.month)
        except Exception:
            current_summary = {
                'year': now.year,
                'month': now.month,
                'income': 0.0,
                'expenses': 0.0,
                'balance': 0.0,
                'transaction_count': 0
            }
        
        # Get last 6 months summaries
        monthly_summaries = []
        for i in range(6):
            month = now.month - i
            year = now.year
            if month <= 0:
                month += 12
                year -= 1
            try:
                summary = db.get_monthly_summary(year, month)
                monthly_summaries.append(summary)
            except Exception:
                monthly_summaries.append({
                    'year': year,
                    'month': month,
                    'income': 0.0,
                    'expenses': 0.0,
                    'balance': 0.0,
                    'transaction_count': 0
                })
        
        # Get top spending categories
        try:
            spending_by_category = db.get_spending_by_category()
            top_categories = spending_by_category[:5] if len(spending_by_category) > 5 else spending_by_category
        except Exception:
            top_categories = []
        
        # Get budget status
        try:
            current_month_str = f"{now.year}-{now.month:02d}"
            budget_status = db.get_budget_status(current_month_str)
        except Exception:
            budget_status = []
        
        if format_type == 'excel':
            return _export_reports_excel(current_summary, monthly_summaries, top_categories, budget_status)
        else:  # PDF (default)
            return _export_reports_pdf(current_summary, monthly_summaries, top_categories, budget_status)
    except Exception as e:
        messages.error(request, f'Error exporting reports: {str(e)}')
        return redirect('tracker:reports')


def _export_reports_pdf(current_summary, monthly_summaries, top_categories, budget_status):
    """Export reports to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Title style
    title_style = ParagraphStyle(
        'Title',
        fontSize=18,
        textColor=colors.HexColor('#2497F9'),
        spaceAfter=20,
        alignment=1
    )
    
    # Section title style
    section_style = ParagraphStyle(
        'Section',
        fontSize=14,
        textColor=colors.HexColor('#2497F9'),
        spaceAfter=10,
        spaceBefore=20
    )
    
    # Title
    title = Paragraph("<b>Mfukoni Finance Tracker - Financial Reports</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Current Month Summary
    elements.append(Paragraph("<b>Current Month Summary</b>", section_style))
    summary_data = [
        ['Metric', 'Value'],
        ['Month/Year', f"{current_summary.get('month_name', current_summary['month'])} {current_summary['year']}"],
        ['Income', f"KES {current_summary['income']:,.2f}"],
        ['Expenses', f"KES {current_summary['expenses']:,.2f}"],
        ['Balance', f"KES {current_summary['balance']:,.2f}"],
        ['Transactions', str(current_summary['transaction_count'])]
    ]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2497F9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        # Make date cell visible (row 1, column 1 - Month/Year value)
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 1), (1, 1), 11),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Monthly Trends
    if monthly_summaries:
        elements.append(Paragraph("<b>6-Month Trend</b>", section_style))
        trend_data = [['Month', 'Income', 'Expenses', 'Balance', 'Transactions']]
        for summary in monthly_summaries:
            trend_data.append([
                f"{summary.get('month_name', summary['month'])} {summary['year']}",
                f"KES {summary['income']:,.2f}",
                f"KES {summary['expenses']:,.2f}",
                f"KES {summary['balance']:,.2f}",
                str(summary['transaction_count'])
            ])
        trend_table = Table(trend_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch])
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2497F9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            # Make date column (column 0) visible - bold black text
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 10),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.black),
        ]))
        elements.append(trend_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Top Categories
    if top_categories:
        elements.append(Paragraph("<b>Top Spending Categories</b>", section_style))
        cat_data = [['Category', 'Total Amount (KES)']]
        for cat in top_categories:
            cat_data.append([cat['category_name'], f"{cat['total']:,.2f}"])
        cat_table = Table(cat_data, colWidths=[2.5*inch, 1.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2497F9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Budget Status
    if budget_status:
        elements.append(Paragraph("<b>Budget Status</b>", section_style))
        budget_data = [['Category', 'Budget', 'Spent', 'Remaining', '% Used']]
        for budget in budget_status:
            budget_data.append([
                budget['category_name'],
                f"KES {budget['budget_limit']:,.2f}",
                f"KES {budget['spent']:,.2f}",
                f"KES {budget['remaining']:,.2f}",
                f"{budget['percentage_used']:.1f}%"
            ])
        budget_table = Table(budget_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 0.8*inch])
        budget_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2497F9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(budget_table)
    
    doc.build(elements)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="mfukoni_reports.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _export_reports_excel(current_summary, monthly_summaries, top_categories, budget_status):
    """Export reports to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from io import BytesIO
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Current Month Summary Sheet
    ws1 = wb.create_sheet("Current Month Summary")
    ws1.append(['Metric', 'Value'])
    ws1.append(['Month/Year', f"{current_summary.get('month_name', current_summary['month'])} {current_summary['year']}"])
    ws1.append(['Income', current_summary['income']])
    ws1.append(['Expenses', current_summary['expenses']])
    ws1.append(['Balance', current_summary['balance']])
    ws1.append(['Transactions', current_summary['transaction_count']])
    
    # Style summary sheet
    header_fill = PatternFill(start_color="2497F9", end_color="2497F9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    date_font = Font(bold=True, color="000000", size=12)  # Black, bold, visible font for dates
    
    # Header row with smooth outer border
    for col_idx, cell in enumerate(ws1[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        left_style = 'medium' if col_idx == 1 else 'thin'
        right_style = 'medium' if col_idx == 2 else 'thin'
        cell.border = Border(
            left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
            right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
            top=Side(style='medium', color='CCCCCC'),
            bottom=Side(style='thin', color='E0E0E0')
        )
    
    # Style date cell (row 2, column 2 - Month/Year value) to make it visible
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White background
    date_cell = ws1['B2']
    date_cell.fill = white_fill
    date_cell.font = Font(bold=True, color="000000", size=12)  # Black, bold, visible font
    date_cell.number_format = '@'  # Format as text to prevent Excel date formatting
    date_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Style all data rows with smooth borders and visible fonts
    data_font = Font(color="000000", size=11)  # Black font for all data cells
    max_row = ws1.max_row
    for row_idx, row in enumerate(ws1.iter_rows(min_row=2, max_row=max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            # Explicitly set white fill to ensure visibility
            cell.fill = white_fill
            
            # Apply visible black font to all data cells
            if row_idx == 2 and col_idx == 2:  # Date cell already styled above
                pass
            else:
                cell.font = data_font
            
            # Determine border style based on position
            left_style = 'medium' if col_idx == 1 else 'thin'
            right_style = 'medium' if col_idx == 2 else 'thin'
            top_style = 'thin'
            bottom_style = 'medium' if row_idx == max_row else 'thin'
            
            cell.border = Border(
                left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                top=Side(style=top_style, color='E0E0E0'),
                bottom=Side(style=bottom_style, color='CCCCCC' if bottom_style == 'medium' else 'E0E0E0')
            )
    
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    
    # Remove unused border variables (they were for reference only)
    
    # Monthly Trends Sheet
    if monthly_summaries:
        ws2 = wb.create_sheet("Monthly Trends")
        ws2.append(['Month', 'Income', 'Expenses', 'Balance', 'Transactions'])
        for summary in monthly_summaries:
            ws2.append([
                f"{summary.get('month_name', summary['month'])} {summary['year']}",
                summary['income'],
                summary['expenses'],
                summary['balance'],
                summary['transaction_count']
            ])
        
        # Header row with smooth outer border
        for col_idx, cell in enumerate(ws2[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            left_style = 'medium' if col_idx == 1 else 'thin'
            right_style = 'medium' if col_idx == 5 else 'thin'
            cell.border = Border(
                left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                top=Side(style='medium', color='CCCCCC'),
                bottom=Side(style='thin', color='E0E0E0')
            )
        
        column_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 12}
        for col, width in column_widths.items():
            ws2.column_dimensions[col].width = width
        
        # Format numbers and style date column with smooth borders and visible fonts
        no_fill = PatternFill(fill_type=None)  # No fill - transparent background
        date_font = Font(bold=True, color="000000", size=11)  # Black, bold, visible font for dates
        data_font = Font(color="000000", size=11)  # Black font for other data cells
        max_row = ws2.max_row
        for row_idx, row in enumerate(ws2.iter_rows(min_row=2, max_row=max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                # Explicitly set no fill to ensure visibility
                cell.fill = no_fill
                
                if col_idx == 1:  # Date column (Month)
                    cell.font = date_font
                    cell.number_format = '@'  # Format as text to prevent Excel date formatting
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx in [2, 3, 4]:  # Income, Expenses, Balance columns
                    cell.font = data_font
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 5:  # Transactions column
                    cell.font = data_font
                
                # Smooth border styling based on position
                left_style = 'medium' if col_idx == 1 else 'thin'
                right_style = 'medium' if col_idx == 5 else 'thin'
                top_style = 'thin'
                bottom_style = 'medium' if row_idx == max_row else 'thin'
                
                cell.border = Border(
                    left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                    right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                    top=Side(style=top_style, color='E0E0E0'),
                    bottom=Side(style=bottom_style, color='CCCCCC' if bottom_style == 'medium' else 'E0E0E0')
                )
    
    # Top Categories Sheet
    if top_categories:
        ws3 = wb.create_sheet("Top Categories")
        ws3.append(['Category', 'Total Amount (KES)'])
        for cat in top_categories:
            ws3.append([cat['category_name'], cat['total']])
        
        # Header row with smooth outer border
        for col_idx, cell in enumerate(ws3[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            left_style = 'medium' if col_idx == 1 else 'thin'
            right_style = 'medium' if col_idx == 2 else 'thin'
            cell.border = Border(
                left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                top=Side(style='medium', color='CCCCCC'),
                bottom=Side(style='thin', color='E0E0E0')
            )
        
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20
        
        # Data rows with smooth borders and visible fonts
        no_fill = PatternFill(fill_type=None)  # No fill - transparent background
        data_font = Font(color="000000", size=11)  # Black font for all data cells
        max_row = ws3.max_row
        for row_idx, row in enumerate(ws3.iter_rows(min_row=2, max_row=max_row), start=2):
            row[1].number_format = '#,##0.00'
            row[1].alignment = Alignment(horizontal='right')
            for col_idx, cell in enumerate(row, start=1):
                # Explicitly set no fill to ensure visibility
                cell.fill = no_fill
                # Apply visible black font to all cells
                cell.font = data_font
                
                left_style = 'medium' if col_idx == 1 else 'thin'
                right_style = 'medium' if col_idx == 2 else 'thin'
                top_style = 'thin'
                bottom_style = 'medium' if row_idx == max_row else 'thin'
                
                cell.border = Border(
                    left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                    right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                    top=Side(style=top_style, color='E0E0E0'),
                    bottom=Side(style=bottom_style, color='CCCCCC' if bottom_style == 'medium' else 'E0E0E0')
                )
    
    # Budget Status Sheet
    if budget_status:
        ws4 = wb.create_sheet("Budget Status")
        ws4.append(['Category', 'Budget', 'Spent', 'Remaining', '% Used'])
        for budget in budget_status:
            ws4.append([
                budget['category_name'],
                budget['budget_limit'],
                budget['spent'],
                budget['remaining'],
                budget['percentage_used']
            ])
        
        # Header row with smooth outer border
        for col_idx, cell in enumerate(ws4[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            left_style = 'medium' if col_idx == 1 else 'thin'
            right_style = 'medium' if col_idx == 5 else 'thin'
            cell.border = Border(
                left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                top=Side(style='medium', color='CCCCCC'),
                bottom=Side(style='thin', color='E0E0E0')
            )
        
        column_widths = {'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 10}
        for col, width in column_widths.items():
            ws4.column_dimensions[col].width = width
        
        # Data rows with smooth borders and visible fonts
        no_fill = PatternFill(fill_type=None)  # No fill - transparent background
        data_font = Font(color="000000", size=11)  # Black font for all data cells
        max_row = ws4.max_row
        for row_idx, row in enumerate(ws4.iter_rows(min_row=2, max_row=max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                # Explicitly set no fill to ensure visibility
                cell.fill = no_fill
                # Apply visible black font to all cells
                cell.font = data_font
                
                if col_idx in [2, 3, 4]:  # Budget, Spent, Remaining columns
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 5:  # Percentage column
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal='right')
                
                # Smooth border styling based on position
                left_style = 'medium' if col_idx == 1 else 'thin'
                right_style = 'medium' if col_idx == 5 else 'thin'
                top_style = 'thin'
                bottom_style = 'medium' if row_idx == max_row else 'thin'
                
                cell.border = Border(
                    left=Side(style=left_style, color='CCCCCC' if left_style == 'medium' else 'E0E0E0'),
                    right=Side(style=right_style, color='CCCCCC' if right_style == 'medium' else 'E0E0E0'),
                    top=Side(style=top_style, color='E0E0E0'),
                    bottom=Side(style=bottom_style, color='CCCCCC' if bottom_style == 'medium' else 'E0E0E0')
                )
    
    buffer = BytesIO()
    wb.save(buffer)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mfukoni_reports.xlsx"'
    return response
